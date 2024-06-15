from .exceptions import GetErrorPass
import openpyxl.workbook
import os
import random

def strong_pass(length=25):
    lower_case = 'abcdefghijklmnopqrstuvwxyz'
    upper_case = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    numbers = '0123456789'
    symbols = '!|^$&£-_()][}@#*{'
    all_characters = upper_case + lower_case + numbers + symbols
    
    if length < 10:
        raise GetErrorPass("La lunghezza della password deve essere almeno 10 caratteri")
    
    return ''.join(random.sample(all_characters * 50, length))

def get_user_data():
    platform = input("Inserisci la piattafarma per cui vuoi registrare l'account: ")
    username = input("Inserisci username da usare: ")
    lenght_pass =  int(input("Inserisci la lunghezza della password che vuoi creare (Un minimo di 10 caratteri): "))
    pasword = strong_pass(lenght_pass)
    
    return {
        "platform" : platform,
        "username" : username,
        "pasword" : pasword
    }
    
def save_user_data_to_file(data,filename="account_data.txt"):
        with open(filename,'a') as file:
            for key, value in data.items():
                file.write(f"{key}: {value}\n")
    
def main():
    user_data = get_user_data()
    save_user_data_to_file(user_data)
    print(f"I tuoi dati sono stati salvati correttamente in  account_data_data.txt")    
    
def create_file_excel():
    if not os.path.exists('platforms_account.xlsx'):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = "Platform"
        sheet['B1'] = "Username"
        sheet['C1'] = "Password"
        wb.save('platforms_account.xlsx')
      
def write_to_excel():
    create_file_excel()
    data = get_user_data()
    wb = openpyxl.load_workbook('platforms_account.xlsx', data_only=True)
    sheet = wb.active
    last_row = sheet.max_row
    
    last_row += 1
    for col_num, key in enumerate(data.keys(), 1):
        sheet.cell(row=last_row, column=col_num, value=data[key])
    
    wb.save('platforms_account.xlsx')
    read_excel()

def read_excel():
    try:
        wb = openpyxl.load_workbook('platforms_account.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
            print(row)
    except FileNotFoundError:
        print("File excel non trovato")

    
    
    
    

